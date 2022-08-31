from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
wb=load_workbook('test.xlsx')
ws=wb.active
name=ws['A3'].value
print(name)